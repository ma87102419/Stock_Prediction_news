#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from openpyxl import load_workbook


# In[2]:


data = pd.read_excel("台灣50三天.xlsx")


# In[35]:


data.head(10)


# In[3]:


datalist = []
for i in range(247):
    datalist.append([data["年月日"][i], data["收盤價(元)"][i]])


# In[4]:


big_num = 0
small_num = 0
for i in range(242):
    for k in range(1,4):
        if datalist[i][1] < datalist[i+k][1]:
            big_num += 1
        if datalist[i][1] > datalist[i+k][1]:
            small_num += 1
    if small_num > big_num:
        datalist[i].append("0")
    elif small_num < big_num:
        datalist[i].append("1")
    else:
        datalist[i].append("0.5")
    big_num = 0
    small_num = 0


# In[5]:


datalist[0:10]


# In[6]:


result = load_workbook(filename = "台灣50三天.xlsx")


# In[7]:


result_sheet = result["工作表1"]
for i in range(242):
    result_sheet.cell(row=i+2, column=3).value = datalist[i][2]


# In[8]:


result.save(filename = "台灣50三天.xlsx")

