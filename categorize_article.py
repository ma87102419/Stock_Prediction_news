#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from openpyxl import load_workbook
import datetime


# In[3]:


data = pd.read_excel("台灣50三天.xlsx")
forum = pd.read_excel("forum.xlsm")


# In[4]:


news = pd.read_excel("news.xlsx")


# In[5]:


up = []
down = []


# In[6]:


data.head(10)


# In[7]:


forum.head(10)


# In[8]:


# data date 變化
date = list(data["年月日"])


for i in range(0, len(date)):
    date[i] = datetime.datetime.strftime(date[i], '%Y/%m/%d')
print(date[4])


# In[8]:


print(data["年月日"][1])
print(forum["post_time"][3])
forum["title"][3]


# In[10]:


for i in range(242):
    if data["漲跌"][i] == 1:
        for k in range(9411):
            if data["年月日"][i].to_period("D") == forum["post_time"][k].to_period('D'):
                up.append([forum["date"][k], forum["title"][k], forum["content"][k]])
    elif data["漲跌"][i] == 0:
        for k in range(9411):
            if data["年月日"][i].to_period("D") == forum["post_time"][k].to_period('D'):
                down.append([forum["date"][k], forum["title"][k], forum["content"][k]])


# In[12]:


print(data["年月日"][2].to_period("D"))
for k in range(9411):
    if data["年月日"][2].to_period("D") == forum["post_time"][k].to_period('D'):
        print(data["年月日"][2].to_period("D"))
        print(k)
    


# In[ ]:


if data["年月日"][0].to_period('D') == forum["post_time"][22].to_period('D'):
    print("true")
else:
    print("false")
print(type(data["年月日"][0]))
print(forum["post_time"][22])
print(data["年月日"][0])


# In[ ]:


print(down[:10])


# In[ ]:


# forum
for i in range(242):
    if data["漲跌"][i] == 1:
        for k in range(9411):
            if data["年月日"][i].to_period("D") == forum["date"][k].to_period('D'):
                up.append([forum["date"][k], forum["title"][k], forum["content"][k]])
    elif data["漲跌"][i] == 0:
        for k in range(9411):
            if data["年月日"][i].to_period("D") == forum["date"][k].to_period('D'):
                down.append([forum["date"][k], forum["title"][k], forum["content"][k]])


# In[12]:


# news
for i in range(242):
    if data["漲跌"][i] == 1:
        for k in range(len(news)):
            if date[i] == news["post_date"][k]:
                up.append([news["post_date"][k], news["title"][k], news["content"][k]])
    elif data["漲跌"][i] == 0:
        for k in range(len(news)):
            if date[i] == news["post_date"][k]:
                down.append([news["post_date"][k], news["title"][k], news["content"][k]])
print(down[:10])


# In[17]:


doutput = open("down三天_output.csv","w",encoding = "utf-8")

doutput.write('date\title\content\n')
for i in range(len(down)):
    for j in range(len(down[i])):
        doutput.write(str(down[i][j]))
        doutput.write('\t')
    doutput.write('\n')
doutput.close()


# In[19]:


up[:5]


# In[22]:


uoutput = open("up三天_output.xlsx","w",encoding = "utf-8")

uoutput.write('date\title\content\n')
for i in range(len(up)):
    for j in range(len(up[i])):
        uoutput.write(str(up[i][j]))
        uoutput.write('\t')
    uoutput.write('\n')
uoutput.close()


# In[25]:



name=['date','title','content']
test=pd.DataFrame(columns=name,data=up)
#print(test)
test.to_csv('testcsv.csv',encoding='utf-8')


# In[26]:


name=['date','title','content']
test=pd.DataFrame(columns=name,data=down)
#print(test)
test.to_csv('downcsv.csv',encoding='utf-8')


# In[23]:


result = load_workbook(filename = "up2三天_output.xlsx")
result_sheet = result["工作表1"]
for i in range(len(up)):
    result_sheet.cell(row=i+2, column=1).value = up[i][0]
    result_sheet.cell(row=i+2, column=2).value = up[i][1]
    result_sheet.cell(row=i+2, column=3).value = up[i][2]

result.save(filename = "up2三天_output.xlsx")


# In[25]:


result = load_workbook(filename = "down2三天_output.xlsx")
result_sheet = result["工作表1"]
for i in range(len(down)):
    result_sheet.cell(row=i+2, column=1).value = down[i][0]
    result_sheet.cell(row=i+2, column=2).value = down[i][1]
    result_sheet.cell(row=i+2, column=3).value = down[i][2]

result.save(filename = "down2三天_output.xlsx")


# In[ ]:




